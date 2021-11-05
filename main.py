# Created by: Jessica Gallo
# Created: 12/20/2020
# Last Modified: 1/07/2020 - gui done
# Restarted: 11/3/2021 - input to excel
# Last ModifiedL 11/5/2021 

# Description:
# This is a python program to allow input of the person, the date, start time and end time and it will
# # calculate total hours worked and save it to an excel file in local machine

'''
TO DO:
- make name selection with combo box (dropdown) to select names or just be able to write
    and store a new name
- save newly created workbook in new workbook section
- add new excel workbook to combo box permanantely
'''

import wx
import wx.adv  # DatePickerCtrl and TimePickerCtrl
from datetime import datetime, date
import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
import os

filename = 'TimeSheetNames.csv'
dataSetCSV = pd.read_csv(filename)
dataset = pd.DataFrame(dataSetCSV)


class TimeSheet(wx.Frame):

    def __init__(self, parent, title):
        super(TimeSheet, self).__init__(parent, title=title, size=(300, 350))

        # sets background color
        self.SetBackgroundColour('#A3E4D7')
        # centers program on screen
        self.Center()

        panel = wx.Panel(self)

        hbox = wx.BoxSizer(wx.HORIZONTAL)

        fgs = wx.FlexGridSizer(9, 2, 9, 30)

        # ==========================================================================================================
        # TEXT |
        # =====
        # Current excel file
        self.excelFile = wx.StaticText(panel, label='Current Excel File:')
        # Current Sheet
        # self.sheet = wx.StaticText(panel, label='Current Sheet:')
        # Date Text
        self.date = wx.StaticText(panel, label='Date:')
        # First Name Text
        self.firstName = wx.StaticText(panel, label='First Name:')
        # Last Name Text
        self.lastName = wx.StaticText(panel, label='Last Name:')
        # Start Text
        self.start = wx.StaticText(panel, label='Start Time:')
        # End Text
        self.end = wx.StaticText(panel, label="End Time:")
        # Submit Button
        self.subbtn = wx.Button(panel, label='Submit')
        self.subbtn.Bind(wx.EVT_BUTTON, self.onSubmit)
        # Review Button
        self.revbtn = wx.Button(panel, label='Open Excel File')
        self.revbtn.Bind(wx.EVT_BUTTON, self.onReview)
        # Total Text
        self.total = wx.StaticText(panel, label='Total Hours Worked:')

        # ==========================================================================================================
        # INPUTS |
        # =======
        # Current Excel File
        self.excelFileList = ['TimeSheet2.xlsx']
        self.cbExcelFile = wx.ComboBox(panel, choices=self.excelFileList, style=wx.CB_SORT)
        self.cbExcelFile.Bind(wx.EVT_COMBOBOX, self.OnSelectExcel, self.cbExcelFile)
        # Current Sheet Manipulating
        # self.sheetList = ['sheet1', 'sheet2']
        # self.cbSheet = wx.ComboBox(panel, choices=self.sheetList, style=wx.CB_SORT)
        # self.cbSheet.Bind(wx.EVT_COMBOBOX, self.OnSelectSheet, self.cbSheet)
        # Date Input
        self.dpcDate = wx.adv.DatePickerCtrl(panel, wx.adv.DP_DROPDOWN)
        self.Bind(wx.adv.EVT_DATE_CHANGED, self.OnDateChanged, self.dpcDate)
        # First Name Input
        self.firstNameList = ['First Name']
        self.cbFirstName = wx.ComboBox(panel, choices=self.firstNameList, style=wx.CB_SORT)
        self.cbFirstName.Bind(wx.EVT_COMBOBOX, self.OnSelectFirstName, self.cbFirstName)
        # Last Name Input
        self.lastNameList = ['Last Name']
        self.cbLastName = wx.ComboBox(panel, choices=self.lastNameList, style=wx.CB_SORT)
        self.cbLastName.Bind(wx.EVT_COMBOBOX, self.OnSelectLastName, self.cbLastName)
        # Start Time Input
        self.tpcStart = wx.adv.TimePickerCtrl(panel)
        self.Bind(wx.adv.EVT_TIME_CHANGED, self.OnTimeInsert, self.tpcStart)
        # End Time Input
        self.tpcEnd = wx.adv.TimePickerCtrl(panel)
        self.Bind(wx.adv.EVT_TIME_CHANGED, self.OnTimeInsert, self.tpcEnd)
        # Totals Hours Worked Input
        self.tcTotal = wx.TextCtrl(panel, style=wx.TE_READONLY)

        # ==========================================================================================================
        # FORMATTING |
        # ===========
        # Formatting columns and rows of the GUI
        fgs.AddMany([self.excelFile, (self.cbExcelFile, 1, wx.EXPAND),  # self.sheet, (self.cbSheet, 1, wx.EXPAND),
                     self.date, (self.dpcDate, 1, wx.EXPAND), self.firstName, (self.cbFirstName, 1, wx.EXPAND),
                     self.lastName, (self.cbLastName, 1, wx.EXPAND), self.start, (self.tpcStart, 1, wx.EXPAND),
                     self.end, (self.tpcEnd, 1, wx.EXPAND), self.subbtn, (self.revbtn, 1, wx.ALIGN_RIGHT),
                     self.total, (self.tcTotal, 1, wx.EXPAND)])

        # Makes the input fields expand when window is expanded
        # fgs.AddGrowableRow(2, 1)
        # fgs.AddGrowableCol(1, 1)

        hbox.Add(fgs, proportion=1, flag=wx.ALL | wx.EXPAND, border=15)
        panel.SetSizer(hbox)

    # ==============================================================================================================
    # BUTTON EVENTS |
    # ==============

    def OnSelectExcel(self, evt):
        # gets input values in current excel file combo box
        self.excelSelection = self.cbExcelFile.GetValue()
        # if box is NULL
        if self.excelSelection is '':
            wx.MessageBox('ERROR: Please enter an excel file to write to.', 'ERROR', wx.OK | wx.ICON_ERROR)
        # if selected or inputed an excel file that exists in the combobox already
        elif self.excelSelection in self.excelFileList:
            print('Found excel list item: ', self.excelSelection)

        # if inputed new excel file name that does not exist in the combobox drop down (to create a new excel file)
        elif self.excelSelection not in self.excelFileList:
            print('New excel file inputed: ', self.excelSelection)

            # create a new excel file

    '''
    def OnSelectSheet(self, evt):
        # gets inputed value in current sheet combo box
        self.sheetSelection = self.cbSheet.GetValue()
        # if box is NULL
        if self.sheetSelection is '':
            wx.MessageBox('ERROR: Please enter a sheet.', 'ERROR', wx.OK | wx.ICON_ERROR)
        # if selected or inputed an sheet that exists in the combobox already
        elif self.sheetSelection in self.sheetList:
            print('Found sheet list item: ', self.sheetSelection)
        # if inputed new sheet name that does not exist in the combobox drop down (to create a new sheet)
        elif self.sheetSelection not in self.sheetList:
            print('New sheet inputed.', self.sheetSelection)
            # create a new sheet in the excel file selected
            # write the basis of the file
            # save the new sheet
            # add sheet name to list permanently
    '''
    def OnDateChanged(self, evt):
        self.sel_date = self.dpcDate.GetValue()
        self.sel_date.Format('%m/%d/%y')
        self.sel_date = str(self.sel_date) # for excel formating

    def OnSelectFirstName(self, evt):
        # FIRST NAME
        # gets value
        self.fistNameSelection = self.cbFirstName.GetValue()

        # checks if name is null (no value)
        if self.fistNameSelection is '':
            wx.MessageBox('ERROR: Please enter a first name.', 'ERROR', wx.OK | wx.ICON_ERROR)
        else:
            # checks if name is on the list, if not it adds it to the list
            if self.fistNameSelection in self.firstNameList:
                # add names to csv
                print(self.fistNameSelection)
                #return self.fistNameSelection
            elif self.fistNameSelection not in self.firstNameList:
                # adds name to combobox
                self.cbFirstName.Append(self.fistNameSelection)

                print(self.fistNameSelection)
                print(self.firstNameList)

    def OnSelectLastName(self, evt):
        # LAST NAME
        self.lastNameSelection = self.cbLastName.GetValue()

        # checks if name is null (no value)
        if self.lastNameSelection is '':
            wx.MessageBox('ERROR: Please enter a last name.', 'ERROR', wx.OK | wx.ICON_ERROR)
        else:
            # checks if name is on the list, if not it adds it to the list
            if self.lastNameSelection in self.lastNameList:
                # csv
                print(self.lastNameSelection)
            elif self.lastNameSelection not in self.lastNameList:
                # adds name to combobox
                self.cbLastName.Append(self.lastNameSelection)

                print(self.lastNameSelection)
                print(self.lastNameList)

    def OnTimeInsert(self, evt):
        # these are strings
        self.startStringTime = ('%s:%s:%s' % self.tpcStart.GetTime())
        self.endStringTime = ('%s:%s:%s' % self.tpcEnd.GetTime())

        # convert start time string to int to subtract and insert into excel
        self.StartTime_str = self.startStringTime
        self.StartTime_object = datetime.strptime(self.StartTime_str, '%H:%M:%S').time()

        # convert end time string to int to subtract and insert into excel
        self.EndTime_str = self.endStringTime
        self.EndTime_object = datetime.strptime(self.EndTime_str, '%H:%M:%S').time()

        # subtracts start time from end time to get total hours worked
        self.totalHoursDay = datetime.combine(date.min, self.EndTime_object) - datetime.combine(date.min, self.StartTime_object)
        self.totalHoursDay = str(self.totalHoursDay)  # string for excel

    def onReview(self, evt):
        print('Opening excel file...')
        #os.system('start EXCEL.EXE TimeSheet2.xlsx')
        os.startfile(self.excelSelection)

    def onSubmit(self, evt):
        confirm = wx.MessageBox('Are you sure you want to input these changes? \n\n Please review your changes.',
                                'Confirm', wx.ICON_EXCLAMATION | wx.YES_NO)
        if confirm == wx.YES:
            print('=====================================')
            # Getting input values
            self.OnSelectExcel(evt)
            # self.OnSelectSheet(evt)
            self.OnDateChanged(evt)
            self.OnSelectFirstName(evt)
            self.OnSelectLastName(evt)
            self.OnTimeInsert(evt)

            self.tcTotal.SetValue(str(self.totalHoursDay))

            # INSERT INTO EXCEL
            if self.excelSelection not in self.excelFileList:
                # makes a new workbook using inputted name
                # user must input .xlsx file extension
                workbook = Workbook()
                sheet = workbook.active

                sheet['A1'] = 'Date'
                sheet['B1'] = 'First Name'
                sheet['C1'] = 'Last Name'
                sheet['D1'] = 'Start Time'
                sheet['E1'] = 'End Time'
                sheet['F1'] = 'Total Hours'

                input = [(self.sel_date, self.fistNameSelection, self.lastNameSelection, self.startStringTime,
                          self.endStringTime, self.totalHoursDay)]

                for row in input:
                    # sheet.append([row]) # iterates over multiple rows/1 column
                    sheet.append(row)

                workbook.save(filename=self.excelSelection)

            elif self.excelSelection in self.excelFileList:
                # opens spreadsheet
                workbook = load_workbook(filename=self.excelSelection)
                sheet = workbook.active

                input = [(self.sel_date, self.fistNameSelection, self.lastNameSelection, self.startStringTime,
                         self.endStringTime, self.totalHoursDay)]

                for row in input:
                    # sheet.append([row]) # iterates over multiple rows/1 column
                    sheet.append(row)

                workbook.save(filename=self.excelSelection)


def main():

    app = wx.App()
    ex = TimeSheet(None, title='Time Sheets')
    ex.Show()
    app.MainLoop()


if __name__ == '__main__':
    main()

