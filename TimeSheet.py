# Created by: Jessica Gallo
# Created: 12/20/2020
# Last Modified: 1/07/2020

# Description:
# This is a python program to allow input of the person, the date, start time and end time and it will
# # calculate total hours worked and save it to an excel file in local machine

'''
TO DO:
- make name selection with combo box (dropdown) to select names or just be able to write
    and store a new name
- time should be set already as 00:00
- date should be set already as MM/DD/YY AND OR calander box
- calculate button to put total hours worked for that day in another box AND put all
    informaiton in excel file
- warning box every time yo uclick calculate to make sure you want to input that data
-
'''

import wx  # wxpython (GUI)
import wx.adv  # wxpython GUI for DatePickerCtrl and TimePickerCtrl
from datetime import datetime, date
# import xlswt
# from xlwt import Workbook
import pandas as pd
import xlsxwriter

'''
# from openpyxl import load_workbook
from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet['A1'] = ''
sheet['B1'] = 'DATE'
sheet['C1'] = 'DATE'
'''
filename = '../../Documents/TimesheetNames.csv'
names = ['Fist Name', 'Last Name']
dataSetCSV = pd.read_csv(filename, names=names)
dataset = pd.DataFrame(dataSetCSV)

# print(dataset.head(10))

array = dataset.values
X = array[:, 0]
Y = array[:, 1]


class TimeSheet(wx.Frame):

    def __init__(self, parent, title):
        super(TimeSheet, self).__init__(parent, title=title, size=(300, 350))

        # sets background color
        self.SetBackgroundColour('#A3E4D7')
        # centers program on screen
        self.Center()

        panel = wx.Panel(self)

        hbox = wx.BoxSizer(wx.HORIZONTAL)

        fgs = wx.FlexGridSizer(9, 2, 9, 30)

        # ==========================================================================================================
        # TEXT |
        # =====
        # Current excel file
        self.excelFile = wx.StaticText(panel, label='Current Excel File:')
        # Current Sheet
        self.sheet = wx.StaticText(panel, label='Current Sheet:')
        # Date Text
        self.date = wx.StaticText(panel, label='Date:')
        # First Name Text
        self.firstName = wx.StaticText(panel, label='First Name:')
        # Last Name Text
        self.lastName = wx.StaticText(panel, label='Last Name:')
        # Start Text
        self.start = wx.StaticText(panel, label='Start Time:')
        # End Text
        self.end = wx.StaticText(panel, label="End Time:")
        # Submit Button
        self.subbtn = wx.Button(panel, label='Submit')
        self.subbtn.Bind(wx.EVT_BUTTON, self.onSubmit)
        # Review Button
        self.revbtn = wx.Button(panel, label='Review')
        self.revbtn.Bind(wx.EVT_BUTTON, self.onReview)
        # Total Text
        self.total = wx.StaticText(panel, label='Total Hours Worked:')

        # ==========================================================================================================
        # INPUTS |
        # =======
        # Current Excel File
        self.excelFileList = ['TimeSheet.xlsx']
        self.cbExcelFile = wx.ComboBox(panel, choices=self.excelFileList, style=wx.CB_SORT)
        self.cbExcelFile.Bind(wx.EVT_COMBOBOX, self.OnSelectExcel, self.cbExcelFile)
        # Current Sheet Manipulating
        self.sheetList = ['sheet1', 'sheet2']
        self.cbSheet = wx.ComboBox(panel, choices=self.sheetList, style=wx.CB_SORT)
        self.cbSheet.Bind(wx.EVT_COMBOBOX, self.OnSelectSheet, self.cbSheet)
        # Date Input
        self.dpcDate = wx.adv.DatePickerCtrl(panel)
        self.Bind(wx.adv.EVT_DATE_CHANGED, self.OnDateChanged, self.dpcDate)
        # First Name Input
        self.firstNameList = Y
        self.cbFirstName = wx.ComboBox(panel, choices=self.firstNameList, style=wx.CB_SORT)
        self.cbFirstName.Bind(wx.EVT_COMBOBOX, self.OnSelectFirstName, self.cbFirstName)
        # Last Name Input
        self.lastNameList = X
        self.cbLastName = wx.ComboBox(panel, choices=self.lastNameList, style=wx.CB_SORT)
        self.cbLastName.Bind(wx.EVT_COMBOBOX, self.OnSelectLastName, self.cbLastName)
        # Start Time Input
        self.tpcStart = wx.adv.TimePickerCtrl(panel)
        self.Bind(wx.adv.EVT_TIME_CHANGED, self.OnTimeInsert, self.tpcStart)
        # End Time Input
        self.tpcEnd = wx.adv.TimePickerCtrl(panel)
        self.Bind(wx.adv.EVT_TIME_CHANGED, self.OnTimeInsert, self.tpcEnd)
        # Totals Hours Worked Input
        self.tcTotal = wx.TextCtrl(panel)

        # ==========================================================================================================
        # FORMATTING |
        # ===========
        # Formatting columns and rows of the GUI
        fgs.AddMany([self.excelFile, (self.cbExcelFile, 1, wx.EXPAND), self.sheet, (self.cbSheet, 1, wx.EXPAND),
                     self.date, (self.dpcDate, 1, wx.EXPAND), self.firstName, (self.cbFirstName, 1, wx.EXPAND),
                     self.lastName, (self.cbLastName, 1, wx.EXPAND), self.start, (self.tpcStart, 1, wx.EXPAND),
                     self.end, (self.tpcEnd, 1, wx.EXPAND), self.subbtn, (self.revbtn, 1, wx.ALIGN_RIGHT),
                     self.total, (self.tcTotal, 1, wx.EXPAND)])

        # Makes the input fields expand when window is expanded
        # fgs.AddGrowableRow(2, 1)
        # fgs.AddGrowableCol(1, 1)

        hbox.Add(fgs, proportion=1, flag=wx.ALL | wx.EXPAND, border=15)
        panel.SetSizer(hbox)

    # ==============================================================================================================
    # BUTTON EVENTS |
    # ==============

    def OnSelectExcel(self, evt):
        # gets inputed value in current excel file combo box
        self.excelSelection = self.cbExcelFile.GetValue()
        # if box is NULL
        if self.excelSelection is '':
            wx.MessageBox('ERROR: Please enter an excel file to write to.', 'ERROR', wx.OK | wx.ICON_ERROR)
        # if selected or inputed an excel file that exists in the combobox already
        elif self.excelSelection in self.excelFileList:
            print('Found excel list item: ', self.excelSelection)

            self.workbook = xlsxwriter.Workbook(self.excelSelection)
        # if inputed new excel file name that does not exist in the combobox drop down (to create a new excel file)
        elif self.excelSelection not in self.excelFileList:
            print('New excel file inputed: ', self.excelSelection)
            # create a new excel file
            #book = Workbook()
            #sheet = book.active

            #sheet['A1'] = 'Name'
            # sheet['B1'] = self.sel_date

            # write the basis of the file (eg Name [A], Data [B-H], Total Time [I] (under data input total hours worked)
            # save this new excel file
            # add the new file name to the combo box list permanently

        # Creates an excel file named TimeSheet & specifies writer
        # writer = pd.ExcelWriter('TimeSheets.xlsx', engine='xlsxwriter')

        # write data to file
        # data.to_excel(writer, 'Sheet1')
        # writer.save()

    def OnSelectSheet(self, evt):
        # gets inputed value in current sheet combo box
        self.sheetSelection = self.cbSheet.GetValue()
        # if box is NULL
        if self.sheetSelection is '':
            wx.MessageBox('ERROR: Please enter a sheet.', 'ERROR', wx.OK | wx.ICON_ERROR)
        # if selected or inputed an sheet that exists in the combobox already
        elif self.sheetSelection in self.sheetList:
            print('Found sheet list item: ', self.sheetSelection)
        # if inputed new sheet name that does not exist in the combobox drop down (to create a new sheet)
        elif self.sheetSelection not in self.sheetList:
            print('New sheet inputed.', self.sheetSelection)
            # create a new sheet in the excel file selected
            # write the basis of the file
            # save the new sheet
            # add sheet name to list permanently

    def OnDateChanged(self, evt):
        self.sel_date = self.dpcDate.GetValue()
        print(self.sel_date.Format('%m/%d/%y'))
        # return sel_date.Format('%m/%d/%y')

    def OnSelectFirstName(self, evt):
        # FIRST NAME
        # gets value
        self.fistNameSelection = self.cbFirstName.GetValue()

        # checks if name is null (no value)
        if self.fistNameSelection is '':
            wx.MessageBox('ERROR: Please enter a name.', 'ERROR', wx.OK | wx.ICON_ERROR)
        else:
            # checks if name is on the list, if not it adds it to the list
            if self.fistNameSelection in self.firstNameList:
                print(self.fistNameSelection)
                return self.fistNameSelection
            elif self.fistNameSelection not in self.firstNameList:
                # adds name to list
                self.firstNameList.append(self.fistNameSelection)
                # adds name to combobox
                self.cbFirstName.Append(self.fistNameSelection)
                print(self.fistNameSelection)
                print(self.firstNameList)

    def OnSelectLastName(self, evt):
        # LAST NAME
        self.lastNameSelection = self.cbLastName.GetValue()

        # checks if name is null (no value)
        if self.lastNameSelection is '':
            wx.MessageBox('ERROR: Please enter a name.', 'ERROR', wx.OK | wx.ICON_ERROR)
        else:
            # checks if name is on the list, if not it adds it to the list
            if self.lastNameSelection in self.lastNameList:
                print(self.lastNameSelection)
                return self.lastNameSelection
            elif self.lastNameSelection not in self.lastNameList:
                # adds name to list
                self.lastNameList.append(self.lastNameSelection)
                # adds name to combobox
                self.cbLastName.Append(self.lastNameSelection)
                print(self.lastNameSelection)
                print(self.lastNameList)

    def OnTimeInsert(self, evt):
        # these are strings
        self.startStringTime = ('%s:%s:%s' % self.tpcStart.GetTime())
        self.endStringTime = ('%s:%s:%s' % self.tpcEnd.GetTime())

        # convert start time string to int to subtract and insert into excel
        self.StartTime_str = self.startStringTime
        self.StartTime_object = datetime.strptime(self.StartTime_str, '%H:%M:%S').time()

        # convert end time string to int to subtract and insert into excel
        self.EndTime_str = self.endStringTime
        self.EndTime_object = datetime.strptime(self.EndTime_str, '%H:%M:%S').time()

        print(self.StartTime_object, '\n', self.EndTime_object)

        # subtracts start time from end time to get total hours worked
        self.totalHoursDay = datetime.combine(date.min, self.EndTime_object) - datetime.combine(date.min, self.StartTime_object)
        print(self.totalHoursDay)
        # return startStringTime, endStringTime

    def onSubmit(self, evt):
        confirm = wx.MessageBox('Are you sure you want to input these changes? \n\n Please review your changes.',
                                'Confirm', wx.ICON_EXCLAMATION | wx.YES_NO)
        if confirm == wx.YES:
            print('=====================================')
            # Getting input values
            self.OnSelectExcel(evt)
            self.OnSelectSheet(evt)
            self.OnDateChanged(evt)
            self.OnSelectFirstName(evt)
            self.OnSelectLastName(evt)
            self.OnTimeInsert(evt)

            self.worksheet = self.workbook.add_worksheet()
            self.worksheet.write(0, 3, self.sel_date)

            # taking those values and putting them to excel on this button click
            self.firstNameEx, self.lastNameEx, self.dateEx, self.totalHoursEx = [self.fistNameSelection,
                                                                                 self.lastNameSelection, self.sel_date,
                                                                                 self.totalHoursDay]


            if dataset[dataset['Last Name'].values == self.lastNameEx]:
                print('Found last name: ', self.lastNameEx)
                if dataset[dataset['First Name'].values == self.firstNameEx]:
                    print('Found first name: ', self.firstNameEx)

        elif confirm == wx.NO:
            return None

    def onReview(self, evt):
        print('Open excel file')


def main():

    app = wx.App()
    ex = TimeSheet(None, title='Time Sheets')
    ex.Show()
    app.MainLoop()


if __name__ == '__main__':
    main()

